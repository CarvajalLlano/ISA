import streamlit as st
import re
import unicodedata
import io
from openpyxl import load_workbook
from openpyxl.styles import Font

# --- CONFIGURACIÓN DE INTERFAZ ---
st.set_page_config(page_title="Liquidador Colvanes 2026", layout="wide")
st.title("🚚 Conciliador de Fletes (MERCANCÍA / PAQUETE / DOCUMENTO)")

# --- TUS CONSTANTES ORIGINALES ---
HOJA_PEDIDOS = "MERCANCIA"
HOJA_TARIFAS = "DEFINITIVO 026"
COSTO_MANEJO_POR_UNIDAD = 6041
UMBRAL_DECLARADO_SEGURO = 750000
PORC_DESCUENTO = 0.25
PORC_SEGURO = 0.005
UMBRAL_SIMILAR = 5000

# --- TUS FUNCIONES DE UTILIDAD (SIN CAMBIOS) ---
def normalizar(texto):
    if texto is None: return ""
    texto = str(texto).strip().upper()
    return unicodedata.normalize("NFKD", texto).encode("ascii", "ignore").decode("ascii")

def limpiar_ciudad(texto):
    t = normalizar(texto)
    if not t: return ""
    if t.strip() == "SANTA FE DE BOGOTA": t = "BOGOTA"
    ciudad, depto = t, ""
    m = re.match(r"^(.*?)\s*\((.*?)\)\s*$", t)
    if m:
        ciudad = (m.group(1) or "").strip()
        depto = (m.group(2) or "").strip()
    else:
        for sep in [" - ", "-", " / ", "/", " , ", ","]:
            if sep in t:
                partes = [p.strip() for p in t.split(sep, 1)]
                if len(partes) == 2: ciudad, depto = partes[0], partes[1]
                break
    def limpiar_ruido(s: str) -> str:
        if not s: return ""
        s = s.replace("DISTRITO CAPITAL", " ").replace("DISTRITO ESPECIAL", " ").replace("D.C.", " ").replace("D C", " ").replace("DC", " ").replace("D.E.", " ").replace("D E", " ").replace("DE", " ")
        return re.sub(r"\s+", " ", s).strip()
    ciudad, depto = limpiar_ruido(ciudad), limpiar_ruido(depto)
    if ciudad == "BOGOTA" or ciudad.startswith("BOGOTA "): return "BOGOTA"
    return f"{ciudad}-{depto}" if depto else ciudad

def _as_int(x):
    if isinstance(x, (int, float)): return int(x)
    try:
        s = str(x or "0").strip().replace(".", "").replace(",", "").replace(" ", "")
        return int(float(s))
    except: return 0

def construir_indices(ws):
    origenes, destinos = {}, {}
    for col in range(1, ws.max_column + 1):
        val = ws.cell(1, col).value
        if val: origenes[limpiar_ciudad(val)] = col
    for row in range(2, ws.max_row + 1):
        val = ws.cell(row, 2).value
        if val: destinos[limpiar_ciudad(val)] = row
    return origenes, destinos

# --- TUS DICCIONARIOS DE TARIFAS ---
TARIFAS_PAQUETERIA = {
    "URBANO":[{"min":1,"max":3,"tarifa":5467},{"min":4,"max":5,"tarifa":8971},{"min":6,"max":8,"tarifa":12143}],
    "REGIONAL":[{"min":1,"max":3,"tarifa":7794},{"min":4,"max":5,"tarifa":11146},{"min":6,"max":8,"tarifa":14770}],
    "NACIONAL":[{"min":1,"max":3,"tarifa":9878},{"min":4,"max":5,"tarifa":14500},{"min":6,"max":8,"tarifa":17761}],
    "REEXPEDIDO":[{"min":1,"max":3,"tarifa":27428},{"min":4,"max":5,"tarifa":36490},{"min":6,"max":8,"tarifa":45067}]
}

TARIFAS_DOCUMENTO = {
    ("DE", "URBANO"): [{"valor": 3862, "adic": 0}],
    ("RF", "URBANO"): [{"valor": 7794, "adic": 1557}],
    ("DE", "NACIONAL"): [{"valor": 8956, "adic": 0}, {"valor": 5619, "adic": 0}, {"valor": 28757, "adic": 0}, {"valor": 13171, "adic": 0}, {"valor": 23219, "adic": 0}],
    ("RF", "NACIONAL"): [{"valor": 17308, "adic": 3196}],
    ("DE", "REEXPEDIDO"): [{"valor": 27428, "adic": 0},{"valor": 23219, "adic": 0}],
    ("DE", "REGIONAL"): [{"valor": 5012, "adic": 0}, {"valor": 7370, "adic": 0}],
    ("RF", "REGIONAL"): [{"valor": 10222, "adic": 0}],
    ("RF", "REEXPEDIDO"): [{"valor": 27428, "adic": 0}],
}

# --- LÓGICA DE PROCESAMIENTO (ADAPTADA PARA RECIBIR OBJETOS) ---

def procesar_todo(file_pedidos, file_t):
    wb_pedidos = load_workbook(file_pedidos)
    wb_tarifas = load_workbook(file_t, data_only=True)

    # --- 1. PROCESAMIENTO DE MERCANCÍA ---
    if HOJA_PEDIDOS in wb_pedidos.sheetnames:
        ws_pedidos = wb_pedidos[HOJA_PEDIDOS]
        ws_tarifas = wb_tarifas[HOJA_TARIFAS]
        
        origenes, destinos = construir_indices(ws_tarifas)
        headers = {str(ws_pedidos.cell(1, c).value).strip().upper(): c for c in range(1, ws_pedidos.max_column + 1) if ws_pedidos.cell(1, c).value}
        
        base_col = ws_pedidos.max_column
        # Agregamos encabezados de resultados
        nuevos_titulos = ["CIUDAD_ORIGEN", "CIUDAD_DESTINO", "VALOR_KILO", "PREFAC_FLETE", "OBS_MATCH", "COMPARA_TOTAL"]
        for i, tit in enumerate(nuevos_titulos, 1):
            cell = ws_pedidos.cell(1, base_col + i)
            cell.value = tit
            cell.font = Font(bold=True)

        for row in range(2, ws_pedidos.max_row + 1):
            # Extracción de datos con limpieza
            orig_raw = ws_pedidos.cell(row, headers.get("ORIGEN", 0)).value
            dest_raw = ws_pedidos.cell(row, headers.get("DESTINO", 0)).value
            
            orig = limpiar_ciudad(orig_raw)
            dest = limpiar_ciudad(dest_raw)
            
            peso = _as_int(ws_pedidos.cell(row, headers.get("PESO FACTURADO", 0)).value)
            unid = _as_int(ws_pedidos.cell(row, headers.get("UNIDADES", 0)).value)
            decl = _as_int(ws_pedidos.cell(row, headers.get("DECLARADO", 0)).value)
            total_ext = _as_int(ws_pedidos.cell(row, headers.get("TOTAL", 0)).value)

            col_o, fila_d = origenes.get(orig), destinos.get(dest)
            
            if col_o and fila_d:
                vk = _as_int(ws_tarifas.cell(fila_d, col_o).value)
                
                if vk > 0:
                    flete_base = peso * vk
                    
                    # --- LÓGICA DE DESCUENTO CONDICIONAL (25%) ---
                    # No aplica descuento si el destino contiene "REEXPEDIDO"
                    es_reexpedido = "REEXPEDIDO" in dest
                    
                    if es_reexpedido:
                        flete_con_descuento = flete_base  # Tarifa plena
                        obs_txt = "OK (REEXPEDIDO SIN DESC)"
                    else:
                        flete_con_descuento = flete_base * (1 - PORC_DESCUENTO)
                        obs_txt = "OK"

                    # Cálculo de componentes adicionales
                    costo_manejo = unid * COSTO_MANEJO_POR_UNIDAD
                    costo_seguro = int(decl * PORC_SEGURO) if decl > UMBRAL_DECLARADO_SEGURO else 0
                    
                    total_calc = int(flete_con_descuento + costo_manejo + costo_seguro)

                    # Escribir resultados en las nuevas columnas
                    ws_pedidos.cell(row, base_col + 1).value = orig
                    ws_pedidos.cell(row, base_col + 2).value = dest
                    ws_pedidos.cell(row, base_col + 3).value = vk
                    ws_pedidos.cell(row, base_col + 4).value = total_calc
                    ws_pedidos.cell(row, base_col + 5).value = obs_txt
                    
                    # Comparación con el total del archivo
                    if total_ext:
                        dif = abs(total_calc - total_ext)
                        ws_pedidos.cell(row, base_col + 6).value = "IGUAL" if dif == 0 else ("SIMILAR" if dif <= UMBRAL_SIMILAR else "DIFERENTE")
                else:
                    ws_pedidos.cell(row, base_col + 5).value = "VALOR KILO EN 0"
            else:
                ws_pedidos.cell(row, base_col + 5).value = "CIUDAD NO ENCONTRADA EN MATRIZ"

    # 2. PAQUETE
    if "PAQUETE" in wb_pedidos.sheetnames:
        ws = wb_pedidos["PAQUETE"]
        h = {str(ws.cell(1,c).value).strip().upper():c for c in range(1,ws.max_column+1) if ws.cell(1,c).value}
        bc = ws.max_column
        for row in range(2, ws.max_row + 1):
            trayecto = normalizar(ws.cell(row, h["TRAYECTO"]).value)
            peso = _as_int(ws.cell(row, h["PESO FACTURADO"]).value)
            decl = _as_int(ws.cell(row, h["DECLARADO"]).value)
            tarifa = None
            if trayecto in TARIFAS_PAQUETERIA:
                for esc in TARIFAS_PAQUETERIA[trayecto]:
                    if esc["min"] <= peso <= esc["max"]:
                        tarifa = esc["tarifa"]
                        break
            if tarifa:
                calc = tarifa + (int(decl * 0.01) if decl > 10000 else 0)
                ws.cell(row, bc+2).value = calc
                ws.cell(row, bc+3).value = "IGUAL" if calc == _as_int(ws.cell(row, h["TOTAL"]).value) else "DIFERENTE"

    # 3. DOCUMENTO
    nombre_doc = "DOCUMENTO " if "DOCUMENTO " in wb_pedidos.sheetnames else "DOCUMENTO"
    if nombre_doc in wb_pedidos.sheetnames:
        ws = wb_pedidos[nombre_doc]
        h = {str(ws.cell(1, c).value).strip().upper(): c for c in range(1, ws.max_column + 1)}
        bc = ws.max_column
        for row in range(2, ws.max_row + 1):
            s_limpio = normalizar(ws.cell(row, h["SERVICIO"]).value).replace(".", "").replace(" ", "")
            t_raw = normalizar(ws.cell(row, h["TRAYECTO"]).value)
            t_final = "REEXPEDIDO" if ("REEXPEDIDO" in t_raw or "NOTIFIC" in t_raw) else ("NACIONAL" if "NACIONAL" in t_raw else ("URBANO" if "URBANO" in t_raw else ("REGIONAL" if "REGIONAL" in t_raw else t_raw)))
            
            opciones = TARIFAS_DOCUMENTO.get((s_limpio, t_final))
            if opciones:
                peso = _as_int(ws.cell(row, h["PESO"]).value)
                total_f = _as_int(ws.cell(row, h["TOTAL"]).value)
                calc = opciones[0]["valor"] + (max(0, peso - 1) * opciones[0].get("adic", 0))
                ws.cell(row, bc+1).value = calc
                ws.cell(row, bc+2).value = "IGUAL" if int(calc) == total_f else "DIFERENTE"

    output = io.BytesIO()
    wb_pedidos.save(output)
    return output.getvalue()

# --- CARGA DE ARCHIVOS EN STREAMLIT ---
file_t = st.file_uploader("Subir Matriz de Tarifas (Valor kilo...)", type="xlsx")
file_p = st.file_uploader("Subir Archivo de Pedidos (ISA...)", type="xlsx")

if file_t and file_p:
    if st.button("🚀 PROCESAR"):
        res = procesar_todo(file_p, file_t)
        st.success("✅ ¡Listo!")
        st.download_button("📥 Descargar Resultado", data=res, file_name="CONCILIADO.xlsx")
